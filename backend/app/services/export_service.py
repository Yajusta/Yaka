"""Service pour l'export des cartes en CSV et Excel."""

import csv
import io
import re
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session, joinedload

from ..models import Card, CardItem


def get_cards_for_export(db: Session) -> List[Card]:
    """
    Récupère toutes les cartes non archivées triées par position de liste puis position de carte.

    Args:
        db: Session de base de données

    Returns:
        Liste des cartes triées
    """
    from ..models import KanbanList

    cards = (
        db.query(Card)
        .filter(Card.is_archived == False)
        .options(
            joinedload(Card.kanban_list),
            joinedload(Card.items),
            joinedload(Card.labels),
            joinedload(Card.assignee),
        )
        .join(KanbanList, Card.list_id == KanbanList.id)
        .order_by(KanbanList.order, Card.position)
        .all()
    )

    return cards


def format_checklist(items: List[CardItem]) -> str:
    """
    Formate la checklist d'une carte.

    Args:
        items: Liste des éléments de checklist

    Returns:
        Chaîne formatée avec [ ] ou [x] et retours chariots
    """
    if not items:
        return ""

    # Trier par position
    sorted_items = sorted(items, key=lambda item: item.position)

    formatted_items = []
    for item in sorted_items:
        checkbox = "[x]" if item.is_done else "[ ]"
        formatted_items.append(f"{checkbox} {item.text}")

    return "\n".join(formatted_items)


def format_labels(card: Card) -> str:
    """
    Formate les étiquettes d'une carte.

    Args:
        card: Carte avec ses étiquettes

    Returns:
        Chaîne avec les étiquettes séparées par " + "
    """
    if not card.labels:
        return ""

    label_names = [label.name for label in card.labels]
    return " + ".join(label_names)


def format_due_date(due_date) -> str:
    """
    Formate la date d'échéance.

    Args:
        due_date: Date d'échéance

    Returns:
        Date au format YYYY-MM-DD ou chaîne vide
    """
    if due_date is None:
        return ""

    if isinstance(due_date, str):
        return due_date

    return due_date.strftime("%Y-%m-%d")


def format_priority(priority) -> str:
    """
    Formate la priorité.

    Args:
        priority: Priorité de la carte

    Returns:
        Nom de la priorité
    """
    if priority is None:
        return ""

    if hasattr(priority, "value"):
        return priority.value

    return str(priority)


def sanitize_csv_text(text: Optional[str]) -> str:
    """
    Nettoie le texte pour le format CSV en remplaçant les retours à la ligne par des espaces.

    Args:
        text: Texte à nettoyer

    Returns:
        Texte nettoyé
    """
    if not text:
        return ""

    # Remplacer tous les types de retours à la ligne par des espaces
    text = text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")

    # Remplacer les espaces multiples par un seul espace
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def generate_csv_export(db: Session) -> bytes:
    """
    Génère un fichier CSV avec toutes les cartes non archivées.

    Note: Pour le CSV, la checklist n'est pas exportée et les retours à la ligne
    sont remplacés par des espaces pour éviter les problèmes de formatage.

    Args:
        db: Session de base de données

    Returns:
        Contenu du fichier CSV en bytes
    """
    cards = get_cards_for_export(db)

    # Créer un buffer en mémoire
    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # En-têtes (sans la colonne Checklist)
    headers = ["Liste", "Titre", "Description", "Etiquettes", "Priorité", "Date d'échéance", "Assigné à"]
    writer.writerow(headers)

    # Données
    for card in cards:
        row = [
            sanitize_csv_text(card.kanban_list.name),
            sanitize_csv_text(card.title),
            sanitize_csv_text(card.description or ""),
            sanitize_csv_text(format_labels(card)),
            format_priority(card.priority),
            format_due_date(card.due_date),
            sanitize_csv_text(card.assignee.display_name if card.assignee else ""),
        ]
        writer.writerow(row)

    # Récupérer le contenu et l'encoder en bytes
    csv_content = output.getvalue()
    output.close()

    return csv_content.encode("utf-8-sig")  # BOM pour Excel


def generate_excel_export(db: Session) -> bytes:
    """
    Génère un fichier Excel avec toutes les cartes non archivées.

    Args:
        db: Session de base de données

    Returns:
        Contenu du fichier Excel en bytes
    """
    cards = get_cards_for_export(db)

    # Créer un workbook
    wb = Workbook()
    ws = wb.active

    # Vérifier que ws n'est pas None (type guard pour pyright)
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    ws.title = "Export Tâches"

    # En-têtes avec style
    headers = ["Liste", "Titre", "Description", "Checklist", "Etiquettes", "Priorité", "Date d'échéance", "Assigné à"]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Données
    for row_num, card in enumerate(cards, start=2):
        ws.cell(row=row_num, column=1, value=card.kanban_list.name)
        ws.cell(row=row_num, column=2, value=card.title)
        ws.cell(row=row_num, column=3, value=card.description or "")
        ws.cell(row=row_num, column=4, value=format_checklist(card.items))
        ws.cell(row=row_num, column=5, value=format_labels(card))
        ws.cell(row=row_num, column=6, value=format_priority(card.priority))
        ws.cell(row=row_num, column=7, value=format_due_date(card.due_date))
        ws.cell(row=row_num, column=8, value=card.assignee.display_name if card.assignee else "")

    # Ajuster la largeur des colonnes
    ws.column_dimensions["A"].width = 20  # Liste
    ws.column_dimensions["B"].width = 30  # Titre
    ws.column_dimensions["C"].width = 40  # Description
    ws.column_dimensions["D"].width = 30  # Checklist
    ws.column_dimensions["E"].width = 20  # Etiquettes
    ws.column_dimensions["F"].width = 12  # Priorité
    ws.column_dimensions["G"].width = 15  # Date d'échéance
    ws.column_dimensions["H"].width = 20  # Assigné à

    # Sauvegarder dans un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


def get_export_filename(format_type: str) -> str:
    """
    Génère le nom de fichier pour l'export.

    Args:
        format_type: Type de format ('csv' ou 'xlsx')

    Returns:
        Nom de fichier formaté
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return f"yaka_export_{timestamp}.{format_type}"
