"""Tests d'intégration pour l'API d'export."""

import csv
import io

import pytest
from openpyxl import load_workbook

from app.models.card import Card, CardPriority
from app.models.card_item import CardItem
from app.models.label import Label
from app.models.user import UserRole
from app.routers.auth import router as auth_router
from app.routers.export import router as export_router


@pytest.mark.asyncio
async def test_export_csv_success(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    create_list_record,
    login_user,
    integration_session_factory,
):
    """Test d'export CSV réussi."""
    seed_admin_user()
    list_id = create_list_record("À faire", 1)
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    # Créer des cartes de test
    db = integration_session_factory()
    try:
        card1 = Card(
            title="Test Card 1",
            description="Description 1",
            list_id=list_id,
            position=0,
            priority=CardPriority.HIGH,
            created_by=2,  # L'utilisateur créé
            is_archived=False,
        )
        db.add(card1)
        db.commit()
    finally:
        db.close()

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/?format=csv",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv; charset=utf-8"
        assert "attachment" in response.headers["content-disposition"]
        assert "yaka_export_" in response.headers["content-disposition"]
        assert ".csv" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_export_csv_content(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    create_list_record,
    login_user,
    integration_session_factory,
):
    """Test le contenu du CSV exporté."""
    seed_admin_user()
    list_id = create_list_record("À faire", 1)
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    # Créer des cartes de test
    db = integration_session_factory()
    try:
        card1 = Card(
            title="Test Card 1",
            description="Description 1",
            list_id=list_id,
            position=0,
            priority=CardPriority.HIGH,
            created_by=2,
            is_archived=False,
        )
        card2 = Card(
            title="Test Card 2",
            description="Description 2",
            list_id=list_id,
            position=1,
            priority=CardPriority.MEDIUM,
            created_by=2,
            is_archived=False,
        )
        db.add_all([card1, card2])
        db.commit()
    finally:
        db.close()

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/?format=csv",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200

        # Lire le contenu CSV
        csv_content = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)

        # Vérifier l'en-tête (sans Checklist)
        assert rows[0] == [
            "Liste",
            "Titre",
            "Description",
            "Etiquettes",
            "Priorité",
            "Date d'échéance",
            "Assigné à",
        ]

        # Vérifier qu'on a 2 cartes
        assert len(rows) == 3  # Header + 2 cartes


@pytest.mark.asyncio
async def test_export_csv_excludes_archived(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    create_list_record,
    login_user,
    integration_session_factory,
):
    """Test que les cartes archivées ne sont pas exportées en CSV."""
    seed_admin_user()
    list_id = create_list_record("À faire", 1)
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    # Créer une carte active et une carte archivée
    db = integration_session_factory()
    try:
        card1 = Card(
            title="Active Card",
            description="Active",
            list_id=list_id,
            position=0,
            priority=CardPriority.HIGH,
            created_by=2,
            is_archived=False,
        )
        card2 = Card(
            title="Archived Card",
            description="Archived",
            list_id=list_id,
            position=1,
            priority=CardPriority.MEDIUM,
            created_by=2,
            is_archived=True,
        )
        db.add_all([card1, card2])
        db.commit()
    finally:
        db.close()

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/?format=csv",
            headers={"Authorization": f"Bearer {token}"},
        )

        csv_content = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)

        # Header + 1 carte active seulement
        assert len(rows) == 2
        assert "Active Card" in rows[1][1]  # Titre dans la 2ème colonne
        assert "Archived Card" not in str(rows)


@pytest.mark.asyncio
async def test_export_excel_success(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    create_list_record,
    login_user,
    integration_session_factory,
):
    """Test d'export Excel réussi."""
    seed_admin_user()
    list_id = create_list_record("À faire", 1)
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    # Créer une carte de test
    db = integration_session_factory()
    try:
        card1 = Card(
            title="Test Card 1",
            description="Description 1",
            list_id=list_id,
            position=0,
            priority=CardPriority.HIGH,
            created_by=2,
            is_archived=False,
        )
        db.add(card1)
        db.commit()
    finally:
        db.close()

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/?format=xlsx",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert "yaka_export_" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_export_excel_content(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    create_list_record,
    login_user,
    integration_session_factory,
):
    """Test le contenu du fichier Excel exporté."""
    seed_admin_user()
    list_id = create_list_record("À faire", 1)
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    # Créer des cartes de test avec checklist
    db = integration_session_factory()
    try:
        card1 = Card(
            title="Test Card 1",
            description="Description 1",
            list_id=list_id,
            position=0,
            priority=CardPriority.HIGH,
            created_by=2,
            is_archived=False,
        )
        db.add(card1)
        db.commit()
        db.refresh(card1)

        # Ajouter checklist
        item1 = CardItem(card_id=card1.id, text="Item 1", is_done=True, position=0)
        item2 = CardItem(card_id=card1.id, text="Item 2", is_done=False, position=1)
        db.add_all([item1, item2])
        db.commit()
    finally:
        db.close()

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/?format=xlsx",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200

        # Charger le workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Vérifier l'en-tête (avec Checklist)
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Liste",
            "Titre",
            "Description",
            "Checklist",
            "Etiquettes",
            "Priorité",
            "Date d'échéance",
            "Assigné à",
        ]

        # Vérifier que la checklist est présente
        assert ws.max_row >= 2  # Header + au moins 1 carte
        checklist_content = ws.cell(row=2, column=4).value
        assert "[x] Item 1" in checklist_content
        assert "[ ] Item 2" in checklist_content


@pytest.mark.asyncio
async def test_export_requires_authentication(async_client_factory):
    """Test que l'export nécessite une authentification."""
    async with async_client_factory(export_router) as client:
        response = await client.get("/export/?format=csv")
        assert response.status_code == 401


@pytest.mark.asyncio
async def test_export_invalid_format(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    login_user,
):
    """Test avec un format invalide."""
    seed_admin_user()
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/?format=pdf",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 400
        assert "Format invalide" in response.json()["detail"]


@pytest.mark.asyncio
async def test_export_missing_format(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    login_user,
):
    """Test sans paramètre format."""
    seed_admin_user()
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 422  # Validation error


@pytest.mark.asyncio
async def test_visitor_can_export(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    create_list_record,
    login_user,
    integration_session_factory,
):
    """Test qu'un visiteur peut exporter."""
    seed_admin_user()
    list_id = create_list_record("À faire", 1)
    create_regular_user("visitor@example.com", "VisitorPass123!", display_name="Visitor", role=UserRole.VISITOR)

    # Créer une carte
    db = integration_session_factory()
    try:
        card1 = Card(
            title="Test Card",
            description="Description",
            list_id=list_id,
            position=0,
            priority=CardPriority.MEDIUM,
            created_by=2,
            is_archived=False,
        )
        db.add(card1)
        db.commit()
    finally:
        db.close()

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "visitor@example.com", "VisitorPass123!")

        response = await client.get(
            "/export/?format=csv",
            headers={"Authorization": f"Bearer {token}"},
        )

        # Les visiteurs peuvent exporter
        assert response.status_code == 200


@pytest.mark.asyncio
async def test_export_empty_database(
    async_client_factory,
    seed_admin_user,
    create_regular_user,
    login_user,
):
    """Test d'export CSV avec une base vide."""
    seed_admin_user()
    create_regular_user("user@example.com", "UserPass123!", display_name="Test User", role=UserRole.EDITOR)

    async with async_client_factory(auth_router, export_router) as client:
        token = await login_user(client, "user@example.com", "UserPass123!")

        response = await client.get(
            "/export/?format=csv",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200

        csv_content = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)

        # Juste l'en-tête, pas de données
        assert len(rows) == 1
