"""Tests unitaires pour le service d'export."""

import csv
import io
import os
import sys
from datetime import date

import pytest
from openpyxl import load_workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models.card import Card, CardPriority
from app.models.card_item import CardItem
from app.models.kanban_list import KanbanList
from app.models.label import Label
from app.models.user import User, UserRole, UserStatus
from app.services import export_service

# Configuration de la base de données de test
TEST_DB_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(TEST_DB_DIR, exist_ok=True)
TEST_DB_PATH = os.path.join(TEST_DB_DIR, "test_export_service.db")
SQLALCHEMY_DATABASE_URL = f"sqlite:///{TEST_DB_PATH}"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@pytest.fixture
def db_session():
    """Fixture pour créer une session de base de données de test."""
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    yield db
    db.close()
    Base.metadata.drop_all(bind=engine)


@pytest.fixture
def sample_user(db_session):
    """Crée un utilisateur de test."""
    user = User(
        email="test@example.com",
        display_name="Test User",
        password_hash="hashed_password",
        role=UserRole.ADMIN,
        status=UserStatus.ACTIVE,
    )
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)
    return user


@pytest.fixture
def sample_lists(db_session):
    """Crée des listes de test."""
    list1 = KanbanList(name="À faire", order=0)
    list2 = KanbanList(name="En cours", order=1)
    list3 = KanbanList(name="Terminé", order=2)
    db_session.add_all([list1, list2, list3])
    db_session.commit()
    db_session.refresh(list1)
    db_session.refresh(list2)
    db_session.refresh(list3)
    return [list1, list2, list3]


@pytest.fixture
def sample_labels(db_session, sample_user):
    """Crée des étiquettes de test."""
    label1 = Label(name="Bug", color="#ff0000", created_by=sample_user.id)
    label2 = Label(name="Feature", color="#00ff00", created_by=sample_user.id)
    db_session.add_all([label1, label2])
    db_session.commit()
    db_session.refresh(label1)
    db_session.refresh(label2)
    return [label1, label2]


@pytest.fixture
def sample_cards(db_session, sample_user, sample_lists, sample_labels):
    """Crée des cartes de test."""
    # Carte 1 : simple
    card1 = Card(
        title="Tâche 1",
        description="Description simple",
        list_id=sample_lists[0].id,
        position=0,
        priority=CardPriority.HIGH,
        created_by=sample_user.id,
        assignee_id=sample_user.id,
        due_date=date(2025, 12, 31),
        is_archived=False,
    )
    card1.labels.append(sample_labels[0])

    # Carte 2 : avec checklist
    card2 = Card(
        title="Tâche 2",
        description="Description avec\nplusieurs lignes\net des retours",
        list_id=sample_lists[1].id,
        position=0,
        priority=CardPriority.MEDIUM,
        created_by=sample_user.id,
        is_archived=False,
    )
    card2.labels.extend(sample_labels)

    # Carte 3 : archivée (ne devrait pas être exportée)
    card3 = Card(
        title="Tâche archivée",
        description="Cette carte est archivée",
        list_id=sample_lists[2].id,
        position=0,
        priority=CardPriority.LOW,
        created_by=sample_user.id,
        is_archived=True,
    )

    db_session.add_all([card1, card2, card3])
    db_session.commit()
    db_session.refresh(card1)
    db_session.refresh(card2)
    db_session.refresh(card3)

    # Ajouter des items de checklist à card2
    item1 = CardItem(card_id=card2.id, text="Item 1", is_done=True, position=0)
    item2 = CardItem(card_id=card2.id, text="Item 2", is_done=False, position=1)
    db_session.add_all([item1, item2])
    db_session.commit()

    return [card1, card2, card3]


class TestFormatFunctions:
    """Tests pour les fonctions de formatage."""

    def test_sanitize_csv_text_with_newlines(self):
        """Test sanitize_csv_text avec des retours à la ligne."""
        text = "Ligne 1\nLigne 2\rLigne 3\r\nLigne 4"
        result = export_service.sanitize_csv_text(text)
        assert result == "Ligne 1 Ligne 2 Ligne 3 Ligne 4"

    def test_sanitize_csv_text_with_none(self):
        """Test sanitize_csv_text avec None."""
        result = export_service.sanitize_csv_text(None)
        assert result == ""

    def test_sanitize_csv_text_with_empty_string(self):
        """Test sanitize_csv_text avec une chaîne vide."""
        result = export_service.sanitize_csv_text("")
        assert result == ""

    def test_sanitize_csv_text_with_multiple_spaces(self):
        """Test sanitize_csv_text avec des espaces multiples."""
        text = "Texte   avec    beaucoup     d'espaces"
        result = export_service.sanitize_csv_text(text)
        assert result == "Texte avec beaucoup d'espaces"

    def test_format_checklist_empty(self):
        """Test format_checklist avec une liste vide."""
        result = export_service.format_checklist([])
        assert result == ""

    def test_format_checklist_with_items(self, db_session):
        """Test format_checklist avec des items."""
        items = [
            CardItem(card_id=1, text="Item 1", is_done=True, position=0),
            CardItem(card_id=1, text="Item 2", is_done=False, position=1),
        ]
        result = export_service.format_checklist(items)
        assert result == "[x] Item 1\n[ ] Item 2"

    def test_format_labels_empty(self, db_session, sample_user):
        """Test format_labels avec une carte sans étiquettes."""
        card = Card(
            title="Test",
            list_id=1,
            position=0,
            priority=CardPriority.MEDIUM,
            created_by=sample_user.id,
        )
        result = export_service.format_labels(card)
        assert result == ""

    def test_format_labels_with_multiple(self, db_session, sample_user, sample_lists, sample_labels):
        """Test format_labels avec plusieurs étiquettes."""
        card = Card(
            title="Test",
            list_id=sample_lists[0].id,
            position=0,
            priority=CardPriority.MEDIUM,
            created_by=sample_user.id,
        )
        card.labels.extend(sample_labels)
        db_session.add(card)
        db_session.commit()
        result = export_service.format_labels(card)
        assert result == "Bug + Feature"

    def test_format_due_date_none(self):
        """Test format_due_date avec None."""
        result = export_service.format_due_date(None)
        assert result == ""

    def test_format_due_date_with_date(self):
        """Test format_due_date avec une date."""
        test_date = date(2025, 12, 31)
        result = export_service.format_due_date(test_date)
        assert result == "2025-12-31"

    def test_format_due_date_with_string(self):
        """Test format_due_date avec une chaîne."""
        result = export_service.format_due_date("2025-12-31")
        assert result == "2025-12-31"

    def test_format_priority(self):
        """Test format_priority."""
        result = export_service.format_priority(CardPriority.HIGH)
        assert result == "high"

    def test_format_priority_none(self):
        """Test format_priority avec None."""
        result = export_service.format_priority(None)
        assert result == ""


class TestGetCardsForExport:
    """Tests pour get_cards_for_export."""

    def test_get_cards_excludes_archived(self, db_session, sample_cards):
        """Test que les cartes archivées sont exclues."""
        cards = export_service.get_cards_for_export(db_session)
        assert len(cards) == 2
        assert all(not card.is_archived for card in cards)

    def test_get_cards_sorted_by_list_and_position(self, db_session, sample_user, sample_lists):
        """Test que les cartes sont triées par liste puis position."""
        # Créer plusieurs cartes dans différentes listes
        card1 = Card(
            title="Card A",
            list_id=sample_lists[1].id,
            position=1,
            priority=CardPriority.MEDIUM,
            created_by=sample_user.id,
        )
        card2 = Card(
            title="Card B",
            list_id=sample_lists[0].id,
            position=0,
            priority=CardPriority.MEDIUM,
            created_by=sample_user.id,
        )
        card3 = Card(
            title="Card C",
            list_id=sample_lists[1].id,
            position=0,
            priority=CardPriority.MEDIUM,
            created_by=sample_user.id,
        )
        db_session.add_all([card1, card2, card3])
        db_session.commit()

        cards = export_service.get_cards_for_export(db_session)
        # Ordre attendu : list[0] (order=0), puis list[1] (order=1)
        # Dans list[1] : position 0 avant position 1
        assert cards[0].title == "Card B"
        assert cards[1].title == "Card C"
        assert cards[2].title == "Card A"


class TestGenerateCSVExport:
    """Tests pour generate_csv_export."""

    def test_csv_export_structure(self, db_session, sample_cards):
        """Test la structure du fichier CSV."""
        csv_bytes = export_service.generate_csv_export(db_session)
        csv_content = csv_bytes.decode("utf-8-sig")

        # Lire le CSV
        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)

        # Vérifier l'en-tête (sans Checklist)
        assert rows[0] == [
            "Liste",
            "Titre",
            "Description",
            "Etiquettes",
            "Priorité",
            "Date d'échéance",
            "Assigné à",
        ]

        # Vérifier qu'on a 2 cartes (pas les archivées)
        assert len(rows) == 3  # Header + 2 cartes

    def test_csv_export_removes_newlines(self, db_session, sample_cards):
        """Test que les retours à la ligne sont supprimés."""
        csv_bytes = export_service.generate_csv_export(db_session)
        csv_content = csv_bytes.decode("utf-8-sig")

        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)

        # Trouver la carte avec description multi-lignes
        for row in rows[1:]:  # Skip header
            if "plusieurs lignes" in row[2]:  # Description est la 3ème colonne
                assert "\n" not in row[2]
                assert "plusieurs lignes et des retours" in row[2]
                break

    def test_csv_export_no_checklist_column(self, db_session, sample_cards):
        """Test que la colonne Checklist n'est pas dans le CSV."""
        csv_bytes = export_service.generate_csv_export(db_session)
        csv_content = csv_bytes.decode("utf-8-sig")

        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)

        assert "Checklist" not in rows[0]


class TestGenerateExcelExport:
    """Tests pour generate_excel_export."""

    def test_excel_export_structure(self, db_session, sample_cards):
        """Test la structure du fichier Excel."""
        excel_bytes = export_service.generate_excel_export(db_session)

        # Charger le workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Vérifier l'en-tête
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Liste",
            "Titre",
            "Description",
            "Checklist",
            "Etiquettes",
            "Priorité",
            "Date d'échéance",
            "Assigné à",
        ]

        # Vérifier qu'on a 2 cartes (pas les archivées)
        assert ws.max_row == 3  # Header + 2 cartes

    def test_excel_export_has_checklist(self, db_session, sample_cards):
        """Test que la checklist est présente dans Excel."""
        excel_bytes = export_service.generate_excel_export(db_session)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Trouver la carte avec checklist
        for row in ws.iter_rows(min_row=2, values_only=True):
            if "Tâche 2" in str(row[1]):  # Titre est la 2ème colonne
                checklist = row[3]  # Checklist est la 4ème colonne
                assert "[x] Item 1" in checklist
                assert "[ ] Item 2" in checklist
                break

    def test_excel_export_preserves_newlines(self, db_session, sample_cards):
        """Test que les retours à la ligne sont préservés dans Excel."""
        excel_bytes = export_service.generate_excel_export(db_session)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Trouver la carte avec description multi-lignes
        for row in ws.iter_rows(min_row=2, values_only=True):
            if "plusieurs lignes" in str(row[2]):  # Description est la 3ème colonne
                assert "\n" in row[2]
                break

    def test_excel_export_column_widths(self, db_session, sample_cards):
        """Test que les largeurs de colonnes sont définies."""
        excel_bytes = export_service.generate_excel_export(db_session)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        assert ws.column_dimensions["A"].width == 20
        assert ws.column_dimensions["B"].width == 30
        assert ws.column_dimensions["C"].width == 40
        assert ws.column_dimensions["D"].width == 30
        assert ws.column_dimensions["E"].width == 20
        assert ws.column_dimensions["F"].width == 12
        assert ws.column_dimensions["G"].width == 15
        assert ws.column_dimensions["H"].width == 20


class TestGetExportFilename:
    """Tests pour get_export_filename."""

    def test_csv_filename_format(self):
        """Test le format du nom de fichier CSV."""
        filename = export_service.get_export_filename("csv")
        assert filename.startswith("yaka_export_")
        assert filename.endswith(".csv")
        assert "_" in filename  # Contient la date et l'heure

    def test_xlsx_filename_format(self):
        """Test le format du nom de fichier Excel."""
        filename = export_service.get_export_filename("xlsx")
        assert filename.startswith("yaka_export_")
        assert filename.endswith(".xlsx")
        assert "_" in filename  # Contient la date et l'heure

    def test_filename_contains_timestamp(self):
        """Test que le nom de fichier contient un timestamp."""
        import re

        filename = export_service.get_export_filename("csv")
        # Format attendu : yaka_export_YYYY-MM-DD_HHMMSS.csv
        pattern = r"yaka_export_\d{4}-\d{2}-\d{2}_\d{6}\.csv"
        assert re.match(pattern, filename)
